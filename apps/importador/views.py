from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.http import require_POST
import openpyxl


@staff_member_required
def vista_importador(request):
    context = {
        "title": "Importar datos desde Excel",
    }
    return render(request, "admin/importador/importador.html", context)


@staff_member_required
@require_POST
def importar_productos(request):
    """Procesa el Excel de productos y carga los datos en la BD."""
    from apps.productos.models import Producto, Proveedor

    archivo = request.FILES.get("archivo")
    if not archivo:
        return JsonResponse({"ok": False, "error": "No se recibio ningun archivo."})

    if not archivo.name.endswith(".xlsx"):
        return JsonResponse({"ok": False, "error": "El archivo debe ser .xlsx"})

    creados = 0
    actualizados = 0
    errores = []

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)

        # Buscar hoja PRODUCTOS
        if "PRODUCTOS" not in wb.sheetnames:
            return JsonResponse({"ok": False, "error": "No se encontro la hoja 'PRODUCTOS' en el archivo."})

        ws = wb["PRODUCTOS"]

        # Encabezados en fila 3, datos desde fila 5
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=3, column=col).value
            if val:
                headers[val.lower().replace(" ", "_")] = col

        for row in range(5, ws.max_row + 1):
            # Si la fila esta vacia, saltar
            codigo_val = ws.cell(row=row, column=headers.get("codigo", 1)).value
            if not codigo_val:
                continue

            try:
                codigo = str(codigo_val).strip()
                nombre = str(ws.cell(row=row, column=headers.get("nombre", 2)).value or "").strip()
                descripcion = str(ws.cell(row=row, column=headers.get("descripcion", 3)).value or "").strip()
                categoria = str(ws.cell(row=row, column=headers.get("categoria", 4)).value or "").strip().upper()
                unidad = str(ws.cell(row=row, column=headers.get("unidad_medida", 5)).value or "UN").strip().upper()
                proveedor_nombre = str(ws.cell(row=row, column=headers.get("proveedor_nombre", 6)).value or "").strip()
                requiere_anmat_val = str(ws.cell(row=row, column=headers.get("requiere_anmat", 8)).value or "NO").strip().upper()
                requiere_anmat = requiere_anmat_val == "SI"
                nro_anmat = str(ws.cell(row=row, column=headers.get("numero_registro_anmat", 9)).value or "").strip()
                stock_min = float(ws.cell(row=row, column=headers.get("stock_minimo_seguridad", 10)).value or 0)
                tiempo_transito = int(ws.cell(row=row, column=headers.get("tiempo_transito_dias", 11)).value or 0)
                tiempo_anmat = int(ws.cell(row=row, column=headers.get("tiempo_tramite_anmat_dias", 12)).value or 0)
                consumo = ws.cell(row=row, column=headers.get("consumo_promedio_mensual", 13)).value
                activo_val = str(ws.cell(row=row, column=headers.get("activo", 14)).value or "SI").strip().upper()
                activo = activo_val != "NO"

                if not nombre or not categoria:
                    errores.append(f"Fila {row}: nombre y categoria son obligatorios.")
                    continue

                # Proveedor
                proveedor = None
                if proveedor_nombre:
                    pais = str(ws.cell(row=row, column=headers.get("pais_origen", 7)).value or "").strip()
                    proveedor, _ = Proveedor.objects.get_or_create(
                        nombre=proveedor_nombre,
                        defaults={"pais_origen": pais or "Sin especificar", "activo": True}
                    )

                defaults = {
                    "nombre": nombre,
                    "descripcion": descripcion,
                    "categoria": categoria,
                    "unidad_medida": unidad,
                    "proveedor": proveedor,
                    "requiere_anmat": requiere_anmat,
                    "numero_registro_anmat": nro_anmat,
                    "stock_minimo_seguridad": stock_min,
                    "tiempo_transito_dias": tiempo_transito,
                    "tiempo_tramite_anmat_dias": tiempo_anmat,
                    "consumo_promedio_mensual": float(consumo) if consumo else None,
                    "consumo_manual": True,
                    "activo": activo,
                }

                producto, created = Producto.objects.update_or_create(
                    codigo=codigo, defaults=defaults
                )
                if created:
                    creados += 1
                else:
                    actualizados += 1

            except Exception as e:
                errores.append(f"Fila {row}: {str(e)}")

    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Error al leer el archivo: {str(e)}"})

    return JsonResponse({
        "ok": True,
        "creados": creados,
        "actualizados": actualizados,
        "errores": errores,
        "total": creados + actualizados,
    })


@staff_member_required
@require_POST
def importar_proveedores(request):
    """Procesa el Excel de proveedores y carga los datos en la BD."""
    from apps.productos.models import Proveedor
    from apps.proveedores.models import FichaProveedor

    archivo = request.FILES.get("archivo")
    if not archivo:
        return JsonResponse({"ok": False, "error": "No se recibio ningun archivo."})

    if not archivo.name.endswith(".xlsx"):
        return JsonResponse({"ok": False, "error": "El archivo debe ser .xlsx"})

    creados = 0
    actualizados = 0
    errores = []

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)

        if "PROVEEDORES" not in wb.sheetnames:
            return JsonResponse({"ok": False, "error": "No se encontro la hoja 'PROVEEDORES' en el archivo."})

        ws = wb["PROVEEDORES"]

        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=3, column=col).value
            if val:
                headers[val.lower().replace(" ", "_")] = col

        for row in range(5, ws.max_row + 1):
            nombre_val = ws.cell(row=row, column=headers.get("nombre", 1)).value
            if not nombre_val:
                continue

            try:
                nombre = str(nombre_val).strip()
                pais = str(ws.cell(row=row, column=headers.get("pais_origen", 2)).value or "").strip()
                activo_val = str(ws.cell(row=row, column=headers.get("activo", 3)).value or "SI").strip().upper()
                activo = activo_val != "NO"

                proveedor, created = Proveedor.objects.update_or_create(
                    nombre=nombre,
                    defaults={"pais_origen": pais or "Sin especificar", "activo": activo}
                )

                # Ficha extendida
                contacto_nombre = str(ws.cell(row=row, column=headers.get("contacto_nombre", 4)).value or "").strip()
                contacto_email = str(ws.cell(row=row, column=headers.get("contacto_email", 5)).value or "").strip()
                contacto_wa = str(ws.cell(row=row, column=headers.get("contacto_whatsapp", 6)).value or "").strip()
                canal = str(ws.cell(row=row, column=headers.get("canal_habitual", 7)).value or "EMAIL").strip().upper()
                moneda = str(ws.cell(row=row, column=headers.get("moneda_habitual", 8)).value or "USD").strip().upper()
                plazo = ws.cell(row=row, column=headers.get("plazo_pago_dias", 9)).value
                modalidad = str(ws.cell(row=row, column=headers.get("modalidad_pago", 10)).value or "").strip()
                notas = str(ws.cell(row=row, column=headers.get("notas", 11)).value or "").strip()

                FichaProveedor.objects.update_or_create(
                    proveedor=proveedor,
                    defaults={
                        "contacto_nombre": contacto_nombre,
                        "contacto_email": contacto_email,
                        "contacto_whatsapp": contacto_wa,
                        "canal_habitual": canal if canal in ["EMAIL", "WHATSAPP", "AMBOS"] else "EMAIL",
                        "moneda_habitual": moneda if moneda in ["USD", "EUR", "CNY", "BRL"] else "USD",
                        "plazo_pago_dias": int(plazo) if plazo else 0,
                        "modalidad_pago": modalidad,
                        "notas": notas,
                    }
                )

                if created:
                    creados += 1
                else:
                    actualizados += 1

            except Exception as e:
                errores.append(f"Fila {row}: {str(e)}")

    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Error al leer el archivo: {str(e)}"})

    return JsonResponse({
        "ok": True,
        "creados": creados,
        "actualizados": actualizados,
        "errores": errores,
        "total": creados + actualizados,
    })


@staff_member_required
def descargar_template(request, tipo):
    """Descarga el template Excel correspondiente."""
    import os
    from django.http import FileResponse, Http404
    from django.conf import settings

    nombres = {
        "productos": "template_productos.xlsx",
        "proveedores": "template_proveedores.xlsx",
    }

    nombre = nombres.get(tipo)
    if not nombre:
        raise Http404

    ruta = os.path.join(settings.BASE_DIR, "static", "templates_excel", nombre)
    if not os.path.exists(ruta):
        raise Http404

    return FileResponse(
        open(ruta, "rb"),
        as_attachment=True,
        filename=nombre,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
